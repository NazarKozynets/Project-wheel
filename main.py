import pyautogui
import pygetwindow as gw
import time
import os
import pyperclip
from PyQt6.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QLabel, QMessageBox, QDialog
from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot, Qt
import sys
from openpyxl import load_workbook
import keyboard

COORDS_FILE = 'coords/button_coords.txt'
LOGIN_BUTTON_FILE = 'coords/login_button_coords.txt'
LOGIN_FIELD_FILE = 'coords/login_field_coords.txt'
PASSWORD_FIELD_FILE = 'coords/password_field_coords.txt'
SECOND_LOGIN_BUTTON_FILE = 'coords/second_login_button_coords.txt'
ERROR_PASSWORD_FIELD_FILE = 'coords/error_password_field_coords.txt'
ERROR_SECOND_LOGIN_BUTTON_FILE = 'coords/error_second_login_button_coords.txt'
MIMIC_WINDOW_TITLE = 'Mimic'
URL_TO_OPEN = 'https://promo.ladbrokes.com/en/promo/bspin/INSTANTSPINS'
EXCEL_FILE = 'excel/users.xlsx'
ERROR_IMAGE = 'screenshots/error_image.png'
ERROR_COORDS_FILE = 'coords/error_coords.txt'


def save_coords(coords, file_name):
    try:
        os.makedirs(os.path.dirname(file_name), exist_ok=True)
        with open(file_name, 'w') as f:
            f.write(f'{coords[0]},{coords[1]}')
            print(f'Координаты сохранены в файл {file_name}: {coords}')
    except Exception as e:
        print(f'Ошибка при сохранении координат в {file_name}: {e}')


def load_coords(file_name):
    if not os.path.exists(file_name):
        print(f'Файл {file_name} не существует.')
        return None
    try:
        with open(file_name, 'r') as f:
            x, y = map(int, f.read().strip().split(','))
            return (x, y)
    except Exception as e:
        print(f'Ошибка при загрузке координат из {file_name}: {e}')
        return None


def click_button(file_name):
    coords = load_coords(file_name)
    if coords:
        print(f'Нажимаем на кнопку или поле по координатам: {coords}')
        pyautogui.click(coords)
    return None


def read_excel(file_name):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        login = row[0] if row[0] else ''
        password = row[1] if row[1] else ''
        column_c = row[2] if row[2] else ''
        column_e = row[4] if row[4] else ''
        users.append({'login': login, 'password': password, 'column_c': column_c, 'column_e': column_e})
    users = [user for user in users if user['login'] and user['password']]
    return users


def write_excel(current_login, type, amount):
    print(f'Сохраняем данные для: {current_login} - {type}: {amount}')
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        target_row = None
        for row in range(1, sheet.max_row + 1):
            login_cell = sheet.cell(row=row, column=1).value
            if login_cell == current_login:
                target_row = row
                break

        if target_row is None:
            print(f'Логин {current_login} не найден в файле')
            return False

        if type == 'pounds':
            sheet.cell(row=target_row, column=3, value=amount)
            print(f'Фунты сохранены: {amount}')
        elif type == 'tokens':
            sheet.cell(row=target_row, column=4, value=amount)
            print(f'Токены сохранены: {amount}')
        else:
            print(f'Неизвестный тип: {type}')
            return False

        wb.save(EXCEL_FILE)
        print(f'Данные успешно сохранены для {current_login}')
        return True

    except Exception as e:
        print(f'Ошибка при работе с Excel: {e}')
        return False


def enter_credentials(login, password, error_coords=False):
    """Ввод данных аккаунта без использования буфера обмена"""

    if not error_coords:
        print('Нажимаем первую кнопку Login')
        click_button(LOGIN_BUTTON_FILE)
        time.sleep(2)

        print(f'Вводим логин: {login}')
        click_button(LOGIN_FIELD_FILE)
        time.sleep(1)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.typewrite(login, interval=0.05)
        time.sleep(0.5)

        print(f'Вводим пароль: {password}')
        click_button(PASSWORD_FIELD_FILE)
        time.sleep(1)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.typewrite(password, interval=0.05)
        time.sleep(1)

        print('Нажимаем вторую кнопку Login')
        click_button(SECOND_LOGIN_BUTTON_FILE)
        time.sleep(2)

    else:
        print('Обрабатываем ввод с ошибкой координат')
        click_button(ERROR_PASSWORD_FIELD_FILE)
        time.sleep(1)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.typewrite(password, interval=0.05)
        time.sleep(1)
        click_button(ERROR_SECOND_LOGIN_BUTTON_FILE)

    print('Ввод завершён.')


def request_new_coords(file_name, cords_name, parent=None):
    """ Функция для сохранения новых кордов """
    if parent:
        QMessageBox.information(parent, f'Настройка координат {cords_name}',
                                f'Наведите курсор на нужную позицию {cords_name} и нажмите F7 для сохранения координат.')

    print('Наведите курсор на нужную позицию и нажмите F7')

    keyboard.wait('f7')

    x, y = pyautogui.position()
    save_coords((x, y), file_name)

    if parent:
        QMessageBox.information(parent, 'Успех', f'Координаты для {cords_name} сохранены в файл {file_name}: {x}, {y}')
    else:
        print(f'Координаты сохранены: {x}, {y}')


def close_modal_window_and_click_wheel():
    krestik_coords = load_coords('coords/krestik.txt')
    time.sleep(3)
    print(f'Нажимаем на крестик по координатам: {krestik_coords}')
    pyautogui.click(krestik_coords)


def open_fullscreen():
    fullscreen_coords = load_coords('coords/fullscreen.txt')
    print(f'Открываем полноэкранный режим по координатам: {fullscreen_coords}')
    pyautogui.click(fullscreen_coords)


def click_first_wheel():
    wheel_coords = load_coords('coords/first_wheel.txt')
    print(f'Нажимаем на первое колесо по координатам: {wheel_coords}')
    pyautogui.doubleClick(wheel_coords)


def click_second_wheel():
    wheel_coords = load_coords('coords/second_wheel.txt')
    print(f'Нажимаем на второе колесо по координатам: {wheel_coords}')
    pyautogui.doubleClick(wheel_coords)


def click_third_wheel():
    wheel_coords = load_coords('coords/third_wheel.txt')
    print(f'Нажимаем на третье колесо по координатам: {wheel_coords}')
    pyautogui.doubleClick(wheel_coords)


def second_click_to_wheel():
    """Выполняет повторный клик по третьему колесу с новыми координатами"""
    third_wheel2_coords = load_coords('coords/third_wheel2.txt')
    print(f'Нажимаем на колесо по новым координатам: {third_wheel2_coords}')
    pyautogui.click(third_wheel2_coords)


def press_enter():
    print('Нажимаем Enter для выполнения поиска.')
    pyautogui.press('enter')
    time.sleep(1)


def press_krestik():
    print('Загружаем координаты для новой точки.')
    target_point5_coords = load_coords('coords/target_point5.txt')
    print(f'Нажимаем на крестик по координатам: {target_point5_coords}')
    pyautogui.click(target_point5_coords)
    time.sleep(1)


def wait_for_mimic_window(timeout=30):
    """Ожидает открытия окна браузера Mimic (упрощенная версия)"""
    print('Ожидаем открытия окна браузера Mimic...')
    start_time = time.time()

    while time.time() - start_time < timeout:
        all_windows = gw.getWindowsWithTitle('')

        for window in all_windows:
            # ВЫВОДИМ ВСЕ СВОЙСТВА ОБЪЕКТА WINDOW
            for attr in dir(window):
                if not attr.startswith('_'):  # Показываем только публичные атрибуты
                    try:
                        value = getattr(window, attr)
                        # Ограничиваем длину вывода для удобства чтения
                        if callable(value):
                            print(f"  {attr}: <method>")
                        else:
                            value_str = str(value)
                            if len(value_str) > 100:
                                value_str = value_str[:100] + "..."
                            print(f"  {attr}: {value_str}")
                    except Exception as e:
                        print(f"  {attr}: <error: {e}>")
            print("=" * 50 + "\n")
            try:
                title = window.title
                is_visible = window.visible
                print(f"Проверяем окно: '{title}'")
                print(f"Видимое: {is_visible}")

                if is_visible and title.strip():
                    if ('mimic' in title.lower() and
                            'multilogin' not in title.lower() and
                            len(title) > 10):
                        try:
                            print(f"Найдено окно Mimic: '{title}', активируем...")
                            window.activate()
                            time.sleep(2)
                            active_window = gw.getActiveWindow()
                            if active_window == window:
                                print(f'Окно браузера Mimic активно: "{title}"')
                                return True
                        except Exception as e:
                            print(f'Ошибка при активации браузера: {e}')
            except Exception as e:
                print(f'Ошибка при проверке окна: {e}')

        print(f'Ожидание браузера... ({int(time.time() - start_time)} сек)')
        time.sleep(1)

    print('Не удалось активировать окно Mimic в течение таймаута.')
    return False


def process_user_account(user):
    enter_credentials(user['login'], user['password'])

    """ Проверка наличия ошибки невероятным путём сравнивания скриншотов """
    success = handle_error_and_retry(user['login'], user['password'])
    time.sleep(2)

    print('Закрываем окно с сохранением пароля.')
    target_point2_coords = load_coords('coords/target_point2.txt')
    pyautogui.click(target_point2_coords)
    time.sleep(3)

    print('Закрываем окно с языком.')
    target_point2_coords = load_coords('coords/target_point2.txt')
    pyautogui.click(target_point2_coords)

    if success:
        print(f"Успешный вход для пользователя: {user['login']}")
    return None


def handle_error_and_retry(login, password):
    print(f'Пытаемся выполнить вход для {login}...')
    retries = 0

    # повторяем попытки пока не исчерпаем лимит
    while retries < 5:
        print('Проверяем наличие ошибки на экране...')
        error_coords = load_coords(ERROR_COORDS_FILE)

        if error_coords:
            try:
                screenshot = pyautogui.screenshot(
                    region=(error_coords[0], error_coords[1], 599, 592)
                )
                screenshot.save('screenshots/screenshot_error_area.png')

                error_location = pyautogui.locate(
                    ERROR_IMAGE, screenshot, confidence=0.8
                )

                if error_location:
                    print(
                        "Ошибка найдена на экране! Пожалуйста, введите новые координаты "
                        "для полей 'Пароль' и 'Второй Login'."
                    )
                    enter_credentials(login, password, error_coords=True)
                    retries += 1
                    time.sleep(3)
                    continue

                # если дошли сюда — ошибки не найдено
                print(f'Успешный вход для {login}. Ошибка не найдена на экране.')
                return True

            except pyautogui.ImageNotFoundException:
                print(
                    'Не удалось найти изображение ошибки, возможно, ошибка отсутствует. '
                    'Продолжаем выполнение.'
                )
                return True

        else:
            # координаты ошибки не загружены/не найдены
            return False

    # если исчерпали retries
    return False


def wait_for_browser_to_close():
    print('Ожидаем закрытия браузера...')
    time.sleep(10)


def find_and_copy_pound():  # Функция поиска и копирования фунтов в DevTools
    for attempt in range(3):

        if attempt > 0:
            pyautogui.press('f12')
            time.sleep(2)

        pyautogui.press('f12')
        time.sleep(3)

        pyautogui.hotkey('ctrl', 'f')
        time.sleep(2)

        pyperclip.copy('£')
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(1)

        pyautogui.press('f3')
        time.sleep(1)
        pyautogui.press('f3')
        time.sleep(1)

        target_point_coords = load_coords('coords/target_point.txt')
        pyautogui.moveTo(target_point_coords)
        pyautogui.doubleClick()
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1)

        copied_data = pyperclip.paste()
        if copied_data:
            time.sleep(1)
            return copied_data
        else:
            print(f'Не удалось скопировать фунты на попытке {attempt + 1}')

    time.sleep(1)
    return None


def find_and_copy_tokens():  # Функция поиска и копирования токенов в DevTools
    for attempt in range(3):
        print(f'Попытка {attempt + 1} из 3 скопировать токены...')

        print('Закрываем DevTools перед началом новой попытки...')
        pyautogui.press('f12')
        time.sleep(1)

        print('Нажимаем клавишу F12.')
        pyautogui.press('f12')
        time.sleep(2)

        print('Нажимаем комбинацию Ctrl + F для поиска.')
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1)

        """ Поиск токенов """

        print('Вводим символ \'coinsbalance\' в строку поиска.')
        pyperclip.copy('coinsbalance')
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(1)

        """ Открытие блока с токенами """

        print('Загружаем координаты для новой точки.')
        target_point3_coords = load_coords('coords/target_point3.txt')
        print(f'Нажимаем на крестик по координатам: {target_point3_coords}')
        pyautogui.click(target_point3_coords)

        """ Копирование токенов """

        print('Загружаем координаты для новой точки.')
        target_point4_coords = load_coords('coords/target_point4.txt')
        print(f'Наводим курсор на точку по координатам: {target_point4_coords}')
        pyautogui.moveTo(target_point4_coords)
        time.sleep(1)
        print('Делаем двойной клик.')
        pyautogui.doubleClick()
        time.sleep(1)
        print('Копируем содержимое с помощью Ctrl + C.')
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1)

        copied_data = pyperclip.paste()
        if copied_data:
            print('Токены успешно скопированы!')
            time.sleep(1)
            return copied_data
        else:
            print(f'Не удалось скопировать токены на попытке {attempt + 1}')

    print('Не удалось скопировать токены после 3 попыток')
    time.sleep(1)
    return None


def wait_if_paused(self):
    if self._is_paused:
        self.msleep(100)
    if not self._is_running:
        pass
    return None


def safe_keyboard_reset():
    """Сброс состояния клавиатуры на всякий случай"""
    try:
        pyautogui.keyUp('ctrl')
        pyautogui.keyUp('shift')
        pyautogui.keyUp('alt')
        pyautogui.keyUp('win')
        print("✓ Сброшено состояние клавиатуры")
    except:
        pass


def main_step(user, selected_wheel="Третье колесо"):
    """ Сборник функций воркера связанных с браузером (от открытия браузера до закрытия) """
    print(f"Обрабатываем пользователя: {user['login']}")
    safe_keyboard_reset()

    print('Запускаем разовый профиль в мультилогине...')
    click_button(COORDS_FILE)
    time.sleep(4)

    """ Проверка открыто ли окно браузера """
    if wait_for_mimic_window():
        """ Открытие полноэкранного режима """
        open_fullscreen()
        time.sleep(8)

        """ Вход в аккаунт """
        process_user_account(user)
        time.sleep(5)

        """ Закрытие модалки """
        close_modal_window_and_click_wheel()
        time.sleep(1)

        """ Клик по выбранному колесу """
        if selected_wheel == 'Первое колесо':
            click_first_wheel()
        elif selected_wheel == 'Второе колесо':
            click_second_wheel()
        elif selected_wheel == 'Третье колесо':
            click_third_wheel()

        """ Прокрутка выбранного колеса """
        time.sleep(3)
        second_click_to_wheel()
        time.sleep(9)

        """ Закрытие модального окна после прокрутки """
        press_krestik()
        time.sleep(1)

        """ Поиск кол-ва фунтов """
        pound_amount = find_and_copy_pound()
        write_excel(user['login'], 'pounds', pound_amount)

        """ Поиск кол-ва токенов """
        tokens_amount = find_and_copy_tokens()
        write_excel(user['login'], 'tokens', tokens_amount)
    else:
        print("Не удалось открыть браузер, пропускаем пользователя")
        return False

    print("Завершаем работу с браузером...")
    pyautogui.hotkey('alt', 'f4')
    time.sleep(1)
    return True


class WorkersThread(QThread):  # Сам воркер - бот который кликает по всему
    update_label = pyqtSignal(str)

    def __init__(self, selected_wheel=None):
        super().__init__()
        self._is_paused = False
        self._is_running = True
        self.selected_wheel = selected_wheel
        self.excel_file = EXCEL_FILE

    def run(self):
        if not self.selected_wheel:
            self.update_label.emit('Колесо не выбрано!')
            return

        self.update_label.emit('Старт процесса...')
        users = read_excel(self.excel_file)
        self.update_label.emit(f'Найдено пользователей: {len(users)}')

        print(users)

        for user in users:
            print(user)
            self.wait_if_paused()
            self.update_label.emit(f'Обработка пользователя: {user["login"]}')

            try:
                main_step(user)
                self.update_label.emit(f'Пользователь {user["login"]} обработан')
            except Exception as e:
                self.update_label.emit(f'Ошибка при обработке {user["login"]}: {e}')
                continue

        self.update_label.emit('Все пользователи обработаны!')

    def wait_if_paused(self):
        while self._is_paused:
            self.msleep(100)

    def resume(self):
        if self._is_paused:
            self._is_paused = False
            print('Поток возобновлен.')


class WheelSelectionWindow(QDialog):  # Окно выбора колеса
    wheel_selected = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выберите колесо')
        self.setFixedSize(400, 200)

        self.setWindowFlags(
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Dialog
        )
        self.setModal(True)

        layout = QVBoxLayout()
        self.first_wheel_button = QPushButton('Первое колесо')
        self.first_wheel_button.clicked.connect(self.select_first_wheel)
        layout.addWidget(self.first_wheel_button)
        self.second_wheel_button = QPushButton('Второе колесо')
        self.second_wheel_button.clicked.connect(self.select_second_wheel)
        layout.addWidget(self.second_wheel_button)
        self.third_wheel_button = QPushButton('Третье колесо')
        self.third_wheel_button.clicked.connect(self.select_third_wheel)
        layout.addWidget(self.third_wheel_button)
        self.setLayout(layout)

    def select_first_wheel(self):
        self.wheel_selected.emit('Первое колесо')
        self.accept()

    def select_second_wheel(self):
        self.wheel_selected.emit('Второе колесо')
        self.accept()

    def select_third_wheel(self):
        self.wheel_selected.emit('Третье колесо')
        self.accept()


class MainWindow(QMainWindow):
    """ Окно программы """
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Ladbrokes')
        self.setFixedSize(400, 200)
        self.selected_wheel = None
        self.excel_file_path = 'excel/users.xlsx'
        self.thread = WorkersThread()
        self.thread.update_label.connect(self.update_status)
        self.init_ui()
        self.setup_global_shortcuts()

    def init_ui(self):
        """Инициализация пользовательского интерфейса"""
        layout = QVBoxLayout()
        self.label = QLabel('Нажмите \'Старт\' для начала работы')
        layout.addWidget(self.label)
        self.start_button = QPushButton('Старт')
        self.start_button.clicked.connect(self.start_process)
        layout.addWidget(self.start_button)
        self.pause_button = QPushButton('Пауза')
        self.pause_button.clicked.connect(self.pause_process)
        layout.addWidget(self.pause_button)
        self.clear_button = QPushButton('Очистить данные')
        self.clear_button.clicked.connect(self.clear_excel_columns)
        layout.addWidget(self.clear_button)
        self.update_coords_button = QPushButton('Обновить координаты кнопок')
        self.update_coords_button.clicked.connect(self.update_button_coordinates)
        layout.addWidget(self.update_coords_button)
        self.select_wheel_button = QPushButton('Выбрать колесо')
        self.select_wheel_button.clicked.connect(self.show_wheel_selection)
        layout.addWidget(self.select_wheel_button)
        self.status_label = QLabel('Можно приступать к работе')
        layout.addWidget(self.status_label)
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def show_wheel_selection(self):
        """Показать окно выбора колеса"""
        self.wheel_selection_window = WheelSelectionWindow()
        self.wheel_selection_window.wheel_selected.connect(self.on_wheel_selected)
        self.wheel_selection_window.show()

    def on_wheel_selected(self, wheel):
        """Обработчик выбора колеса"""
        self.selected_wheel = wheel
        self.label.setText(f'Вы выбрали: {wheel}')

    def execute_wheel_code(self):
        """Выполнение кода для выбранного колеса"""
        if not self.selected_wheel:
            QMessageBox.warning(self, 'Ошибка', 'Колесо не выбрано!')
        return None

    def update_label(self, message):
        """Обновление текста метки"""
        self.label.setText(message)

    @pyqtSlot(str)
    def update_status(self, message):
        """Обновление статуса (слот)"""
        self.label.setText(message)

    def setup_global_shortcuts(self):
        """Настройка глобальных горячих клавиш"""
        keyboard.add_hotkey('f9', self.resume_process)
        keyboard.add_hotkey('f8', self.pause_process)

    def start_process(self):
        """Запуск процесса"""
        if not self.thread.isRunning():
            self.thread = WorkersThread(self.selected_wheel)
            self.thread.update_label.connect(self.update_status)
            self.thread._is_paused = False
            self.thread.start()
        return None

    def pause_process(self):
        if self.thread.isRunning():
            print("Процесс приостановлен")
            self.thread._is_paused = True
            self.thread._is_running = False
            self.thread.update_label.emit('Процесс приостановлен')

    def resume_process(self):
        if self.thread.isRunning():
            print("Процесс возобновлен")
            self.thread._is_paused = False
            self.thread._is_running = True
            self.thread.update_label.emit('Процесс возобновлен')

    def clear_excel_columns(self):
        """Очистка колонок C и D в Excel"""
        if not self.excel_file_path:
            QMessageBox.critical(self, 'Ошибка', 'Файл Excel не указан.')
            return None

        try:
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2):
                if len(row) >= 3:
                    row[2].value = None
                if len(row) >= 4:
                    row[3].value = None

            workbook.save(self.excel_file_path)
            workbook.close()

            QMessageBox.information(self, 'Успех', 'Колонки C и D успешно очищены.')

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось очистить файл Excel:\n{e}')

    def update_button_coordinates(self):
        try:
            # Список всех координат для настройки
            coord_configs = [
                (COORDS_FILE, 'кнопки запуска разового профиля'),
                ('coords/fullscreen.txt', 'кнопки полного окна'),
                (LOGIN_BUTTON_FILE, 'кнопки логина'),
                (LOGIN_FIELD_FILE, 'поля логина'),
                (PASSWORD_FIELD_FILE, 'поля пароля'),
                (SECOND_LOGIN_BUTTON_FILE, 'второй кнопки логина'),
                ('coords/target_point2.txt', 'закрытия модального окна с сохранением пароля'),
                ('coords/krestik.txt', 'крестика модального окна'),
                ('coords/first_wheel.txt', 'первого колеса'),
                ('coords/second_wheel.txt', 'второго колеса'),
                ('coords/third_wheel.txt', 'третьего колеса'),
                ('coords/third_wheel2.txt', 'колеса для прокрутки'),
                ('coords/target_point5.txt', 'крестика после прокрутки'),
                ('coords/target_point.txt', 'фунта'),
                ('coords/target_point3.txt', 'места открытия ледбаксов'),
                ('coords/target_point4.txt', 'количества ледбаксов'),
            ]

            # Настройка всех координат через цикл
            for file_name, description in coord_configs:
                request_new_coords(file_name, description, self)

            QMessageBox.information(self, 'Успех', 'Координаты обновлены и сохранены.')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось обновить координаты: {e}')

    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        keyboard.unhook_all_hotkeys()
        self.thread.terminate() if self.thread.isRunning() else None
        QApplication.quit()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
